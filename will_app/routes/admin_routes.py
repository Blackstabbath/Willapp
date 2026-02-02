from flask import Blueprint, send_file, request, session, redirect, url_for, jsonify
import os
import openpyxl
from datetime import datetime, date
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# 📁 SETUP
# ============================================================
admin_bp = Blueprint('admin', __name__)

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "absolute2024"


# ============================================================
# 🔐 AUTHENTICATION
# ============================================================
def admin_required(f):
    def wrapper(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin.admin_login'))
        return f(*args, **kwargs)
    wrapper.__name__ = f.__name__
    return wrapper


@admin_bp.route('/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        if (request.form.get('username') == ADMIN_USERNAME and
                request.form.get('password') == ADMIN_PASSWORD):
            session['admin_logged_in'] = True
            return redirect(url_for('admin.admin_dashboard'))
        return "<h3 style='text-align:center;margin-top:60px;color:#e11d48;'>Invalid login</h3>"
    return """
    <html><head><title>Admin Login</title></head>
    <body style="margin:0;height:100vh;display:flex;align-items:center;justify-content:center;
    background:linear-gradient(135deg,#eff6ff,#e0f2fe);font-family:'Segoe UI';">
      <form method='POST' style="background:white;padding:40px 45px;border-radius:14px;
      box-shadow:0 8px 30px rgba(0,0,0,.1);width:340px;animation:fade .6s;">
        <h2 style='color:#1e3a8a;text-align:center;margin:0 0 25px;'>🔐 Admin Login</h2>
        <label>Username</label><input name='username' required
         style='width:100%;margin:8px 0 18px;padding:10px;border-radius:8px;border:1px solid #cbd5e1;'>
        <label>Password</label><input type='password' name='password' required
         style='width:100%;margin:8px 0 22px;padding:10px;border-radius:8px;border:1px solid #cbd5e1;'>
        <button style='width:100%;padding:12px;background:#1e3a8a;color:#fff;border:none;
        border-radius:8px;font-weight:600;cursor:pointer;transition:.25s;'>Login</button>
      </form>
    <style>@keyframes fade{from{opacity:0;transform:translateY(8px);}to{opacity:1;}}</style>
    </body></html>
    """


@admin_bp.route('/logout')
def admin_logout():
    session.pop('admin_logged_in', None)
    return redirect(url_for('admin.admin_login'))


# ============================================================
# 📊 DASHBOARD
# ============================================================
def get_submission_stats():
    excel_file = "will_data_log.xlsx"
    stats = {"total": 0, "today": 0, "docs": 0}
    if os.path.exists(excel_file):
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        stats["total"] = max(ws.max_row - 1, 0)
        today = date.today().strftime("%Y-%m-%d")
        stats["today"] = sum(1 for r in ws.iter_rows(min_row=2, values_only=True)
                             if today in str(r[0]))
    folder = "generated_wills"
    if os.path.exists(folder):
        stats["docs"] = len([f for f in os.listdir(folder) if f.endswith(".docx")])
    return stats


@admin_bp.route('/')
@admin_required
def admin_dashboard():
    s = get_submission_stats()
    return f"""
    <html><head><title>Admin Dashboard</title></head>
    <body style="margin:0;font-family:'Segoe UI';background:linear-gradient(135deg,#f0f9ff,#e0f2fe);
    animation:fade .6s;">
      <div style='max-width:1100px;margin:60px auto;padding:0 30px;'>
        <header style='display:flex;justify-content:space-between;align-items:center;margin-bottom:40px;'>
          <h1 style='color:#1e3a8a;'>🏠 Absolute Wills Admin</h1>
          <a href='/admin/logout' style='background:#e11d48;color:#fff;padding:10px 18px;
            border-radius:8px;text-decoration:none;font-weight:600;transition:.25s;'>Logout</a>
        </header>
        <section style='display:grid;grid-template-columns:repeat(auto-fit,minmax(250px,1fr));gap:25px;'>
          <div class='card'><div>{s['total']}</div><span>Total Submissions</span></div>
          <div class='card'><div>{s['today']}</div><span>Today</span></div>
          <div class='card'><div>{s['docs']}</div><span>Documents</span></div>
        </section>
        <nav style='margin-top:45px;display:flex;flex-wrap:wrap;gap:20px;'>
          <a href='/admin/documents' class='btn'>📂 Document Library</a>
          <a href='/admin/export' class='btn'>📊 Export Data</a>
          <a href='/' class='btn gray'>🏠 Return Home</a>
        </nav>
      </div>
      <style>
        @keyframes fade{{from{{opacity:0;transform:translateY(8px);}}to{{opacity:1;}}}}
        .card{{background:white;padding:25px;border-radius:12px;text-align:center;
          box-shadow:0 3px 20px rgba(0,0,0,0.08);transition:.3s;}}
        .card:hover{{transform:translateY(-4px);}}
        .card div{{font-size:2.6em;font-weight:700;color:#1e3a8a;}}
        .card span{{color:#475569;font-weight:600;}}
        .btn{{flex:1;min-width:220px;text-align:center;padding:16px 0;border-radius:10px;
          background:linear-gradient(135deg,#1e3a8a,#2563eb);color:#fff;text-decoration:none;
          font-weight:600;box-shadow:0 5px 15px rgba(37,99,235,0.35);transition:.25s;}}
        .btn:hover{{transform:translateY(-3px);filter:brightness(1.1);}}
        .gray{{background:#475569;}}
      </style>
    </body></html>
    """


# ============================================================
# 📁 DOCUMENT LIBRARY
# ============================================================
@admin_bp.route('/documents')
@admin_required
def list_documents():
    folder = "generated_wills"
    docs = []
    if os.path.exists(folder):
        for f in os.listdir(folder):
            if f.endswith(".docx"):
                docs.append({
                    "name": f,
                    "created": datetime.fromtimestamp(os.path.getctime(os.path.join(folder, f))).strftime("%Y-%m-%d %H:%M:%S")
                })
    docs.sort(key=lambda x: x["created"], reverse=True)
    rows = "".join(
        f"<div class='row fade'><input type='checkbox' value='{d['name']}' class='chk'>"
        f"<span>{d['name']}</span><small>{d['created']}</small></div>"
        for d in docs) or "<p class='empty'>No documents yet.</p>"
    return f"""
    <html><head><title>Documents</title></head>
    <body style="margin:0;font-family:'Segoe UI';background:linear-gradient(135deg,#f0f9ff,#e0f2fe);
    animation:fade .6s;">
      <div style='max-width:950px;margin:60px auto;padding:0 25px;'>
        <header style='display:flex;justify-content:space-between;align-items:center;margin-bottom:30px;'>
          <h2 style='color:#1e3a8a;'>📂 Document Library</h2>
          <a href='/admin' style='text-decoration:none;color:#2563eb;font-weight:600;'>← Back</a>
        </header>
        <div class='toolbar'>
          <label><input type='checkbox' id='selectAll' onchange='toggleAll()'> Select All</label>
          <div>
            <button onclick='downloadSel()'>⬇️ Download</button>
            <button class='del' onclick='deleteSel()'>🗑 Delete</button>
          </div>
        </div>
        <section class='list'>{rows}</section>
      </div>
      <style>
        @keyframes fade{{from{{opacity:0;transform:translateY(8px);}}to{{opacity:1;}}}}
        .toolbar{{display:flex;justify-content:space-between;align-items:center;
          background:white;padding:14px 20px;border-radius:10px;
          box-shadow:0 3px 20px rgba(0,0,0,0.08);margin-bottom:18px;}}
        .toolbar button{{background:#2563eb;color:white;border:none;padding:10px 14px;
          border-radius:8px;font-weight:600;margin-left:8px;cursor:pointer;transition:.25s;}}
        .toolbar button:hover{{transform:scale(1.05);}}
        .del{{background:#dc2626;}}
        .list{{background:white;border-radius:10px;box-shadow:0 3px 20px rgba(0,0,0,0.07);
          padding:12px 18px;}}
        .row{{display:flex;align-items:center;justify-content:space-between;
          border-bottom:1px solid #e5e7eb;padding:8px 0;transition:.25s;}}
        .row:hover{{background:#f1f5f9;transform:translateX(3px);}}
        .row span{{flex:1;color:#1e3a8a;font-weight:600;margin-left:10px;}}
        .empty{{text-align:center;color:#6b7280;padding:40px;}}
        .fade{{animation:fade .5s ease-in;}}
      </style>
      <script>
        function toggleAll() {{
          const all = document.getElementById('selectAll').checked;
          document.querySelectorAll('.chk').forEach(c=>c.checked=all);
        }}
        async function deleteSel() {{
          const f=[...document.querySelectorAll('.chk:checked')].map(c=>c.value);
          if(!f.length) return alert('Select files first');
          if(!confirm('Delete selected files?')) return;
          const r=await fetch('/admin/delete-files',{{
            method:'POST',headers:{{'Content-Type':'application/json'}},body:JSON.stringify({{files:f}})
          }});const j=await r.json();alert(j.message);location.reload();
        }}
        function downloadSel() {{
          const f=[...document.querySelectorAll('.chk:checked')].map(c=>c.value);
          if(!f.length) return alert('Select files first');
          f.forEach(v=>window.open('/admin/download/'+encodeURIComponent(v),'_blank'));
        }}
      </script>
    </body></html>
    """


# ============================================================
# 🗑 DELETE FILES
# ============================================================
@admin_bp.route('/delete-files', methods=['POST'])
@admin_required
def delete_files():
    data = request.get_json()
    files = data.get('files', [])
    folder = "generated_wills"
    count = 0
    for f in files:
        p = os.path.join(folder, f)
        if os.path.exists(p):
            os.remove(p)
            count += 1
    return jsonify({"message": f"Deleted {count} file(s)."})


# ============================================================
# ⬇️ DOWNLOAD FILE
# ============================================================
@admin_bp.route('/download/<filename>')
@admin_required
def download_file(filename):
    path = os.path.join("generated_wills", filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return "<h3 style='color:#e11d48;text-align:center;'>File not found.</h3>", 404


# ============================================================
# 📦 EXPORT DATA (Styled & Professional)
# ============================================================
@admin_bp.route('/export')
@admin_required
def export_data():
    excel_file = "will_data_log.xlsx"
    if not os.path.exists(excel_file):
        return "No data to export", 404

    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active

        # 🟦 Header Styling
        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 🧊 Freeze top row
        ws.freeze_panes = "A2"

        # 🔠 Auto-fit columns width
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 4, 12), 60)

        # 💾 Save styled export
        export_filename = f"will_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(export_filename)
        return send_file(export_filename, as_attachment=True)

    except Exception as e:
        return f"Error exporting data: {e}", 500
